from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.sheet_properties.tabColor = "cc3399" #rgb형태로 값을 넣어주면 탭 색상 변경됨.
ws1=wb.create_sheet("YourSheet")
wb.save("sample.xlsx")
